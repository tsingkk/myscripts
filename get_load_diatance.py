# 对Sheet1中多列数据求值，并将结果输出为Sheet2中一列

import urllib.request  # 发送请求
from urllib import parse  # URL编码
import json  # 解析json数据
from openpyxl import load_workbook  # 从Excel中读取乡镇名称


def city_location(localname, somewhere, thekey):
    url1 = (
        'http://restapi.amap.com/v3/place/text?keywords=' + localname +
        '&city=' + somewhere +
        '&citylimit=true&output=json&offset=1&page=1&key=' + thekey
    )
    # 将一些符号进行URL编码
    newUrl1 = parse.quote(url1, safe="/:=&?#+!$,;'@()*[]")
    # 发送请求
    response1 = urllib.request.urlopen(newUrl1)
    # 读取数据
    data1 = response1.read()
    # 解析json数据
    jsonData1 = json.loads(data1)
    # 从json文件中提取座标
    try:
        a = jsonData1['pois'][0]['location']
    except:
        return '查询不到坐标'
    else:
        return a


def road_long(origin, destination, thekey):
    url2 = (
        'https://restapi.amap.com/v3/direction/driving?origin=' + origin +
        '&destination=' + destination +
        '&extensions=all&strategy=1&output=json&key=' + thekey
    )
    # 将一些符号进行URL编码
    newUrl2 = parse.quote(url2, safe="/:=&?#+!$,;'@()*[]")
    # 发送请求
    response2 = urllib.request.urlopen(newUrl2)
    # 接收数据
    data2 = response2.read()
    # 解析json文件
    jsonData2 = json.loads(data2)
    # 从json文件中提取距离
    try:
        a = jsonData2['route']['paths'][0]['distance']
    except:
        return '查询不到距离'
    else:
        return a

somewhere = input('地点所在市或省名称：')
thekey = input('高德开放平台密钥：')
book = load_workbook('workbook.xlsx')
nameSheet = book["Sheet1"]
locationsheet = book['Sheet2']
tol = 0
for index1 in range(ord('A'), ord('A')+nameSheet.max_column):
    index1 = chr(index1)
    nameList = []
    for row in range(1, nameSheet.max_row+1):
        nameList.append(nameSheet[index1+str(row)].value)
    nameList = [x for x in nameList if x is not None]  # 删除空字符
    print(nameList)

    dict1 = {}  # 创建一个字典用于接收坐标数据
    dict2 = {}  # 创建一个字典用于接收距离数据
    for cityname in nameList:
        dict1[cityname] = city_location(cityname, somewhere, thekey)
    for cityname in nameList[1::]:
        if dict1[cityname] == '查询不到坐标':
            dict2[cityname] = '此地查询不到坐标值'
        else:
            dict2[cityname] = road_long(
                dict1[cityname], dict1[nameSheet[index1+'1'].value], thekey
            )
    dict2[nameSheet[index1+'1'].value] = '行政中心'
    print(dict1)
    print(dict2)

    list1 = []
    for cityname in nameList:
        list2 = [cityname, dict1[cityname], dict2[cityname]]
        list1.append(list2)
    print(list1)
    for row in range(len(nameList)):
        (
            locationsheet['A%d' % (row+1+tol)],
            locationsheet['B%d' % (row+1+tol)],
            locationsheet['C%d' % (row+1+tol)]
        ) = list1[row]
    tol += len(nameList)
book.save('workbook.xlsx')
